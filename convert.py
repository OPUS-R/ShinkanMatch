import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import re
from openpyxl import load_workbook
from collections import OrderedDict

def convert_date(val):
    if pd.isna(val): return ""
    s = str(val).strip()
    match = re.search(r"(?:\d{4}[年/]?)?\s*(\d{1,2})[月/](\d{1,2})", s)
    if match: return f"{int(match.group(1))}/{int(match.group(2))}"
    return s

def convert_time_list(val):
    if pd.isna(val): return []
    times = str(val).replace("，", ",").split(",")
    result = []
    for t in times:
        t = t.strip().translate(str.maketrans("０１２３４５６７８９：－〜～", "0123456789:-~~"))
        if re.fullmatch(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}", t):
            result.append(t)
        elif m := re.match(r"^(\d{1,2})[-~](\d{1,2})$", t):
            result.append(f"{int(m.group(1)):02}:00-{int(m.group(2)):02}:00")
        elif m := re.match(r"^(\d{1,2})時[-~〜](\d{1,2})時$", t):
            result.append(f"{int(m.group(1)):02}:00-{int(m.group(2)):02}:00")
    return result

def col_letter_to_index(letter):
    if not letter: return None
    result = 0
    for c in letter: result = result * 26 + ord(c.upper()) - ord('A') + 1
    return result - 1

def parse_column_ranges(text):
    cols = []
    for part in text.split(","):
        if "-" in part:
            a, b = part.split("-")
            cols.extend(range(col_letter_to_index(a), col_letter_to_index(b)+1))
        else:
            cols.append(col_letter_to_index(part))
    return sorted(set(cols))

def process(file, sheet, name_col, line_col, mail_col, inst_col, date_cols):
    df = pd.read_excel(file, sheet_name=sheet, header=None)

    # 日付行（1行目）の変換
    date_map = {}
    for col in date_cols:
        date = convert_date(df.iloc[0, col])
        if date: date_map[col] = date

    merged = OrderedDict()
    for i in range(1, len(df)):  # 全行処理（1行目は日付行）
        name = df.iloc[i, name_col] if name_col is not None else ""
        if not str(name).strip(): continue  # 名前が空欄ならスキップ
        key = str(name).strip()

        line = df.iloc[i, line_col] if line_col is not None else ""
        mail = df.iloc[i, mail_col] if mail_col is not None else ""
        inst = df.iloc[i, inst_col] if inst_col is not None else ""

        # 備考（A列と指定列・日付列を除外した列から収集、A列＝col0は除く）
        remarks = []
        for j in range(1, df.shape[1]):
            if j not in [name_col, line_col, mail_col, inst_col] + date_cols:
                v = df.iloc[i, j]
                if pd.notna(v): remarks.append(str(v).strip())
        remarks_text = " / ".join(remarks)

        # 初回エントリ
        if key not in merged:
            merged[key] = {
                "メールアドレス": mail,
                "名前": key,
                "LINE名": line,
                "希望楽器": inst,
                "備考": remarks_text,
                "スケジュール": {date: set() for date in date_map.values()}
            }

        person = merged[key]

        # 備考は常に上書き（後の方を優先）
        if remarks_text:
            person["備考"] = remarks_text
        if pd.notna(mail): person["メールアドレス"] = mail
        if pd.notna(line): person["LINE名"] = line
        if pd.notna(inst): person["希望楽器"] = inst

        # 時間のマージ（重複を避ける）
        for col, date in date_map.items():
            times = convert_time_list(df.iloc[i, col])
            person["スケジュール"][date].update(times)

    # 出力行を構築
    all_dates = sorted(list(date_map.values()))
    rows = []
    for person in merged.values():
        row = {
            "メールアドレス": person["メールアドレス"],
            "名前": person["名前"],
            "LINE名": person["LINE名"],
            "希望楽器": person["希望楽器"],
            "備考": person["備考"]
        }
        for date in all_dates:
            times = sorted(person["スケジュール"][date])
            row[date] = ", ".join(times)
        rows.append(row)

    df_out = pd.DataFrame(rows)
    df_out = df_out[["メールアドレス", "名前", "LINE名"] + all_dates + ["希望楽器", "備考"]]
    df_out.insert(0, "", "")  # A列空白

    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
    if save_path:
        df_out.to_excel(save_path, index=False)
        messagebox.showinfo("完了", f"保存しました：\n{save_path}")

def start_gui():
    root = tk.Tk()
    root.title("統合対応：横長Excel整形ツール")

    sheet_names = []

    def browse_file():
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        entry_file.delete(0, tk.END)
        entry_file.insert(0, path)

        try:
            wb = load_workbook(path)
            sheet_names.clear()
            sheet_names.extend(wb.sheetnames)
            dropdown_menu["menu"].delete(0, "end")
            for s in sheet_names:
                dropdown_menu["menu"].add_command(label=s, command=tk._setit(sheet_var, s))
            sheet_var.set(sheet_names[0])
        except Exception as e:
            messagebox.showerror("読み込みエラー", str(e))

    def run():
        try:
            file = entry_file.get()
            sheet = sheet_var.get()
            name_col = col_letter_to_index(entry_name.get())
            line_col = col_letter_to_index(entry_line.get())
            mail_col = col_letter_to_index(entry_mail.get())
            inst_col = col_letter_to_index(entry_inst.get())
            date_cols = parse_column_ranges(entry_date.get())
            process(file, sheet, name_col, line_col, mail_col, inst_col, date_cols)
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    # GUIレイアウト
    tk.Label(root, text="Excelファイル").grid(row=0, column=0)
    entry_file = tk.Entry(root, width=40)
    entry_file.grid(row=0, column=1)
    tk.Button(root, text="参照", command=browse_file).grid(row=0, column=2)

    inputs = [
        ("名前の列（例: B）", "entry_name"),
        ("LINE名の列（例: C）", "entry_line"),
        ("メールの列（例: A）", "entry_mail"),
        ("希望楽器の列（例: D）", "entry_inst"),
        ("日付列範囲（例: E-G,J）", "entry_date")
    ]
    widgets = {}
    for idx, (label, varname) in enumerate(inputs, start=1):
        tk.Label(root, text=label).grid(row=idx, column=0)
        entry = tk.Entry(root)
        entry.grid(row=idx, column=1)
        widgets[varname] = entry

    global entry_name, entry_line, entry_mail, entry_inst, entry_date
    entry_name = widgets["entry_name"]
    entry_line = widgets["entry_line"]
    entry_mail = widgets["entry_mail"]
    entry_inst = widgets["entry_inst"]
    entry_date = widgets["entry_date"]

    tk.Label(root, text="シート選択").grid(row=6, column=0)
    sheet_var = tk.StringVar()
    dropdown_menu = tk.OptionMenu(root, sheet_var, "")
    dropdown_menu.grid(row=6, column=1)

    tk.Button(root, text="実行", command=run, bg="lightgreen").grid(row=7, column=1, pady=10)
    root.mainloop()

start_gui()
